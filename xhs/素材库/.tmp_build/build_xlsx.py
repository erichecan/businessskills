import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

FONT = "Arial"
header_font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
body_font = Font(name=FONT, size=10)
wrap = Alignment(wrap_text=True, vertical="top")

def style_sheet(ws, headers, col_widths):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

# ---------------- Sheet 1: 内容表 ----------------
ws1 = wb.active
ws1.title = "内容表"
headers1 = ["关键词","来源","新增/已收","标题","作者/站点","热度","核心要点","标题套路","链接"]
content = [
("职场表达","网页","已收","职场上，10种沟通技巧你知道哪个？","知乎","—","清单体职场沟通技巧汇总","数字清单","https://zhuanlan.zhihu.com/p/555854764"),
("职场表达","网页","已收","如何练就高薪人士的职场沟通技巧","知乎","—","以高薪人士为标签讲沟通风格差异","身份标签","https://zhuanlan.zhihu.com/p/25560466"),
("职场表达","网页","已收","干货|职场沟通的8个要点","知乎","—","职场沟通8要点清单","数字清单","https://zhuanlan.zhihu.com/p/359639851"),
("职场表达","网页","已收","想在职场成功？16个沟通技巧不可少！","HR人力资源管理案例网","—","16个沟通技巧清单","数字清单+悬念","https://www.hrsee.com/?id=601"),
("职场表达","网页","今日新增","职场上如何避免无效沟通","经理人参阅","—","清晰简洁语言、避免模糊词汇是避免无效沟通关键","疑问句式","https://www.managerreads.com/117087.html"),

("面试技巧","网页","已收","前程无忧发布2026届毕业生秋招避坑指南","前程无忧","—","秋招趋势与避坑建议","悬念缺口(避坑)","https://xz.chsi.com.cn/xz/zyts/202509/20250910/2293415299.html"),
("面试技巧","网页","已收","2026年面试技巧怎么提升？这6个实战要点照着练","知乎","—","4维度评估+结构化答题+提问技巧+语速控制+模拟练习","数字+祈使句","https://zhuanlan.zhihu.com/p/2045593194876887565"),
("面试技巧","网页","已收","想进字节的应届生必看！2026校招面试全流程+避坑技巧","知乎","—","字节校招全流程攻略","身份标签+悬念","https://zhuanlan.zhihu.com/p/2024889492524254560"),
("面试技巧","网页","已收","2026留学生国内求职全攻略：网申渠道+面试技巧","高顿留学生求职","—","留学生求职攻略","身份标签+攻略体","https://www.highmarkcareer.com/EmploymentStatus/16723.html"),
("面试技巧","网页","已收","2026留学生国内求职全攻略：网申渠道+面试技巧","青林职途","—","同源转载内容","身份标签+攻略体","https://www.zhiruihou.com/newsdetails/943.html"),
("面试技巧","网页","已收","基于人性的招聘面试实战技巧","名课堂","—","招聘方视角的面试培训课程","理论标签","https://m.mingketang.com/openCourse/desc?id=737059"),

("面试什么话该说","网页","已收","【就业宝典】5分钟读完28个面试加分技巧","广东财经大学","—","28个加分技巧速读","数字+时间承诺","https://csgg.gdufe.edu.cn/_t27/2021/1205/c1163a151315/page.htm"),
("面试什么话该说","网页","已收","面试时应该问面试官什么问题？","超级简历WonderCV","—","反问环节加分问题示例","疑问句式","https://www.wondercv.com/blog/vm7raryl.html"),
("面试什么话该说","网页","已收","让巧妙提问为面试加分","静安政务","—","巧妙提问技巧","悬念(巧妙)","https://www.jingan.gov.cn/rmtzx/003001/20080916/14324648-5b49-4d66-936f-013432491077.html"),
("面试什么话该说","网页","已收","面试官告知可以问几个问题时如何提问能得到加分","知乎","—","反问加分策略讨论","原句形态(用户提问)","https://www.zhihu.com/question/378319301"),
("面试什么话该说","网页","已收","面试最后可以向面试官提什么问题？","知乎","—","结尾反问建议清单","疑问句式","https://zhuanlan.zhihu.com/p/94141723"),
("面试什么话该说","网页","已收","面试到底应该怎么回答最加分？","简历网","—","加分回答策略","悬念(到底)","https://www.jianli.com/article/gxpqqj.html"),

("面试什么话不能说","网页","已收","小错不犯，大错却不断，面试中不容忽视的禁忌行为","界面新闻","—","面试禁忌行为盘点","对比反差体","https://www.jiemian.com/article/1911581.html"),
("面试什么话不能说","网页","已收","面试忌讳","会计学院","—","面试忌讳清单","直白标题","https://kjxy.fjjxu.edu.cn/2015/0304/c82a1425/page.htm"),
("面试什么话不能说","网页","已收","面试时绝对不能说的13句话","CSDN","—","13句禁忌话清单","数字+绝对化措辞","https://blog.csdn.net/yangyouni/article/details/38679483"),
("面试什么话不能说","网页","已收","面试技巧：切记面试中四大禁忌","华图教育","—","四大禁忌清单","数字清单","https://qh.huatu.com/2024/0708/1752672.html"),
("面试什么话不能说","网页","已收","有哪些话是面试的时候千万不能说的？","知乎","—","用户讨论禁忌话题","原句形态(疑问)","https://www.zhihu.com/question/20055335"),
("面试什么话不能说","网页","已收","面试时一定要避开的十大面试禁忌！","简历本","—","十大禁忌清单","数字清单+感叹号","https://www.jianliben.com/article/detail/43392"),
("面试什么话不能说","网页","已收","面试禁忌：这几句话千万别说！","知乎","—","禁忌话清单","悬念+感叹号","https://zhuanlan.zhihu.com/p/13275593806"),
("面试什么话不能说","网页","已收","面试过程中应注意的问题与禁忌汇总","新浪博客","—","禁忌汇总","汇总体","https://blog.sina.com.cn/s/blog_5f28fe8a0102yb7q.html"),

("涨薪历程","网页","今日新增","分享一下我的工作经历，也是对之前工作的一个总结吧","电鸭社区","—","月薪2500运维工程师起步4个月加薪300后跳槽手游公司的真实经历","原句形态(口语分享体)","https://eleduck.com/posts/GzfKNY"),
("涨薪历程","网页","今日新增","一个资深广告文案的涨薪经历：6年7次","数英网","—","6年涨薪7次资深文案真实经历,第4次直接涨到10K","具体数字+时间跨度叙事体","https://www.digitaling.com/articles/29912.html"),
("涨薪历程","网页","今日新增","分享一次申请涨薪的经历","V2EX","—","上海工作7年余两次晋升三次涨薪每次均涨3K现薪23K","原句形态(论坛分享体)","https://v2ex.com/t/1171756"),
("涨薪历程","网页","此前已收","从外贸小白到独自创业十年经历分享","中国制造网","—","（历史已收录,本轮判定重复未展开）","—","https://service.made-in-china.com/service/successful-story/xeXQnJVOImHg.html"),
("涨薪历程","网页","此前已收","从辍学到IT工程师：一路逆袭的奋斗历程","CSDN","—","（历史已收录,本轮判定重复未展开）","—","https://blog.csdn.net/weixin_39634719/article/details/112258659"),
("涨薪历程","网页","此前已收","涨薪30%！申请涨薪就按这3个步骤来（含话术）","知乎","—","（历史已收录,本轮判定重复未展开）","—","https://zhuanlan.zhihu.com/p/659871236"),
("涨薪历程","网页","此前已收","谈谈我的职场经历与升职感悟","CSDN","—","（历史已收录,本轮判定重复未展开）","—","https://blog.csdn.net/u014316335/article/details/135415955"),

("hr问你上一份工作的薪资怎么回答","网页","已收","面试官问你的期望薪资是多少，如何巧妙回应？","MBA智库资讯","—","巧妙回应期望薪资问题的策略","疑问句式","https://news.mbalib.com/story/255486"),
("hr问你上一份工作的薪资怎么回答","网页","已收","面试官问你期望的工资是多少，怎么回答最好？","CSDN","—","期望薪资回答策略","疑问句式","https://blog.csdn.net/agonie201218/article/details/127610437"),
("hr问你上一份工作的薪资怎么回答","网页","已收","怎么跟HR谈薪资？","知乎","—","谈薪整体策略","疑问句式","https://zhuanlan.zhihu.com/p/57505274"),
("hr问你上一份工作的薪资怎么回答","网页","今日新增","HR问我薪资期望，我怎么回答才好？","知乎","—","提供完整薪资结构展现综合价值而非单一数字","原句形态","https://zhuanlan.zhihu.com/p/672939046"),
("hr问你上一份工作的薪资怎么回答","网页","今日新增","当HR问：你对薪资有什么要求，怎么巧妙回答？","赣州人事人才网","—","巧妙回答薪资要求的技巧","疑问句式","https://www.797rs.com/article/16829.html"),
("hr问你上一份工作的薪资怎么回答","网页","已收","谈薪实战：当HR问你目前的薪资是多少时如何既不撒谎又保留溢价空间","Gank Interview","—","既不撒谎又保留溢价空间的谈薪实战技巧","悬念+实战标签","https://www.gankinterview.cn/blog/salary-negotiation-tactics-when-hr-asks-what-is-your-current-salary-how-to-avoid"),

("谈薪话术怎么争取","网页","已收","面试谈薪话术","喜马拉雅","—","谈薪话术问答", "直白标题","https://m.ximalaya.com/ask/t5733190"),
("谈薪话术怎么争取","网页","今日新增","面试谈薪资的技巧话术你了解嘛？","哔哩哔哩","—","谈薪技巧话术清单","疑问句式","https://www.bilibili.com/read/cv8342887"),
("谈薪话术怎么争取","网页","今日新增","【附话术】职场新人谈薪技巧","哔哩哔哩","—","职场新人谈薪技巧+附话术","身份标签(职场新人)","https://www.bilibili.com/read/cv18847316"),
("谈薪话术怎么争取","网页","已收","万能谈薪话术，薪资直接涨50%","知乎","—","万能谈薪话术","悬念(万能)+数字承诺","https://zhuanlan.zhihu.com/p/607403705"),
("谈薪话术怎么争取","网页","已收","谈薪技巧加薪话术","三茅人力资源网","—","谈薪加薪话术合集","直白标题","https://www.hrloo.com/news/276864.html"),
("谈薪话术怎么争取","网页","今日新增","满分谈薪话术怎么说-跟hr谈薪资有哪些套路和话术","果果圈模板","—","满分谈薪话术+套路总结","悬念(满分)","https://m.ggq.com/study/22552.html"),

("hr说这已经是最高薪资了","网页","已收","干货|面试过了，如何谈薪资？拿高薪Offer有哪些技巧？","博客园","—","面试过后谈薪技巧","疑问句式","https://www.cnblogs.com/hogwarts/p/16148096.html"),
("hr说这已经是最高薪资了","网页","今日新增","HR问：假如公司给不到你期望的薪资怎么办？这个问题该如何体面地回答？","CSDN","—","体面回应公司给不到期望薪资的情况","悬念+疑问句式","https://blog.csdn.net/g6U8W7p06dCO99fQ3/article/details/129395229"),
("hr说这已经是最高薪资了","网页","已收","hr提出薪资后还能再谈吗？","知乎","—","薪资提出后能否再谈的讨论","原句形态(疑问)","https://www.zhihu.com/question/361832937"),
("hr说这已经是最高薪资了","网页","今日新增","还没面试 HR 就问期望薪资，该怎么回？","知乎","—","应对提前问薪资的场景","原句形态(疑问)","https://www.zhihu.com/question/530095146"),
("hr说这已经是最高薪资了","网页","已收","[经验帖]上班当然是为了钱，不要避讳。分享一些和HR谈薪的技巧","V2EX","—","谈薪心态+技巧分享","原句形态(论坛分享体)","https://www.v2ex.com/t/928808"),
]
for row in content:
    ws1.append(row)
style_sheet(ws1, headers1, [26,10,12,42,20,10,42,18,46])
for r in range(2, ws1.max_row+1):
    for c in range(1, len(headers1)+1):
        cell = ws1.cell(row=r, column=c)
        cell.font = body_font
        cell.alignment = wrap

# ---------------- Sheet 2: 根词表现表 ----------------
ws2 = wb.create_sheet("根词表现表")
headers2 = ["根词","类型","本次抓取数","本次新增数","本次命中率","累计运行次数","平均热度","升级/退休","备注"]
perf = [
("职场表达","种子",5,1,"20.0%",214,"约1803(样本量少)","维持种子(永不退休)","web-only,长期高频复投后正常低位表现"),
("面试技巧","种子",6,0,"0.0%",211,"约938","维持种子(永不退休)","6条候选全部已收录,web-only"),
("面试什么话该说","种子",6,0,"0.0%",203,"约14","维持种子(永不退休)","web-only"),
("面试什么话不能说","种子",8,0,"0.0%",202,"约19","维持种子(永不退休)","web-only"),
("涨薪历程","活跃",7,3,"42.9%",2,"约1541","维持活跃","长空档复查(距08-16隔16天),3篇均为真实个人涨薪经历叙事"),
("hr问你上一份工作的薪资怎么回答","活跃",6,2,"33.3%",3,"约14467","维持活跃","长空档复查(距08-16隔16天)"),
("谈薪话术怎么争取","候选",6,3,"50.0%",1,"—(web无点赞数据)","维持候选(标记强候补)","候选首投超40%阈值,按账号惯例单次不够暂不升级,下轮复投确认"),
("hr说这已经是最高薪资了","候选",5,2,"40.0%",1,"—(web无点赞数据)","维持候选(标记强候补)","候选首投恰达40%阈值,下轮复投确认"),
]
for row in perf:
    ws2.append(row)
style_sheet(ws2, headers2, [30,10,12,12,12,12,18,22,40])
for r in range(2, ws2.max_row+1):
    for c in range(1, len(headers2)+1):
        cell = ws2.cell(row=r, column=c)
        cell.font = body_font
        cell.alignment = wrap

# ---------------- Sheet 3: 词库验证表 ----------------
ws3 = wb.create_sheet("词库验证表")
headers3 = ["长句","场景类型","三条件-下拉补全","三条件-笔记数量级","三条件-赞藏水平","判定结果","答案空缺/备注"]
verify = [
("发完offer还能再谈薪么","事件","未执行(小红书未登录)","未执行","未执行","跳过验证-维持候选","已连续排队多轮,累计跳过验证达10次以上,建议下次小红书恢复登录后最高优先级验证"),
("怎么样算口头offer","事件","未执行(小红书未登录)","未执行","未执行","跳过验证-维持候选","同上,与前一词同批次排队"),
("口头说发offer","事件","未执行(小红书未登录)","未执行","未执行","跳过验证-维持候选","同上,与前一词同批次排队"),
("收到offer了还能谈薪资吗","事件","未执行(小红书未登录)","未执行","未执行","跳过验证-维持候选","首次因未登录被跳过,备注列本轮补齐跳过记录"),
("面试官暗示你被录用了","事件","未执行(小红书未登录)","未执行","未执行","跳过验证-维持候选","首次因未登录被跳过,备注列本轮补齐跳过记录"),
]
for row in verify:
    ws3.append(row)
style_sheet(ws3, headers3, [30,10,20,16,16,20,44])
for r in range(2, ws3.max_row+1):
    for c in range(1, len(headers3)+1):
        cell = ws3.cell(row=r, column=c)
        cell.font = body_font
        cell.alignment = wrap

# 追加"新词收割清单"作为词库验证表下方补充区块
start_row = ws3.max_row + 3
ws3.cell(row=start_row, column=1, value="本轮新词收割清单").font = Font(name=FONT, bold=True, size=11)
start_row += 1
harvest_headers = ["类型","新词","来源","备注"]
for i,h in enumerate(harvest_headers,1):
    c = ws3.cell(row=start_row, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
start_row += 1
harvest_rows = [
("根词候选(关键词池)","外企offer为何跳过谈薪环节","WebSearch知乎question/563802359衍生","为什么外企发offer跳过谈薪环节的现象讨论,与既有offer/谈薪簇角度互补"),
("长句候选(词库)","三明治拒绝法真的有用吗","WebSearch知乎question/1933933031627929251衍生","答案空缺=现有内容均讲方法本身,缺方法失效后的进阶应对"),
]
for row in harvest_rows:
    start_row += 1
    for i,v in enumerate(row,1):
        cell = ws3.cell(row=start_row, column=i, value=v)
        cell.font = body_font
        cell.alignment = wrap

note_row = start_row + 3
ws3.cell(row=note_row, column=1, value="说明：本次运行小红书未登录(内置浏览器登录墙+Chrome扩展未连接双重确认)，第3步三条件验证与第4步来源①②③（下拉补全/相关搜索/评论区原话）整体跳过；评论区原话.csv本轮无新增。以上仅为web-only可执行部分的交付结果。").font = Font(name=FONT, italic=True, size=9, color="FF0000")

wb.save("/sessions/sweet-confident-hopper/mnt/素材库/职场表达与面试技巧_2026-09-01.xlsx")
print("saved")
