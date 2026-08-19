# -*- coding: utf-8 -*-
import json, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D = '/sessions/keen-trusting-keller/mnt/素材库/'
TODAY = '2026-08-18'
data = json.load(open(D + '.run_0818r2/excel_data.json'))
rows_out = data['rows']  # 关键词,来源,新增/已收,标题,作者/站点,热度,链接
stat = data['stat']

wb = Workbook()

HEAD_FILL = PatternFill(start_color='FFDCE6F1', end_color='FFDCE6F1', fill_type='solid')
HEAD_FONT = Font(bold=True)
NEW_FILL = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'

# ---- 核心要点 / 标题套路 简易推断（按关键词规则化标注） ----
POINT_HINTS = {
 '职场表达': '沟通话术/接话技巧/汇报表达',
 '面试技巧': '综合面试建议/自我介绍/临场应对',
 '面试什么话该说': '英文/中文面试标准表达模板',
 '面试什么话不能说': '面试雷区禁忌句式清单',
 '实习谈薪': '实习薪资谈判流程与话术',
 '涨薪案例': '真实涨薪经历与复盘',
 '推迟面试话术': '改期/推迟面试沟通模板',
 '会议礼仪发言': '会议发言/上台礼仪技巧',
 '面试改期': '改期沟通话术',
}
TITLE_PATTERN_HINTS = {
 '职场表达': '能力动词化标题（"XX的重要性"）/清单体数字公式',
 '面试技巧': '潜台词解码体/清单体万能话术',
 '面试什么话该说': '英文场景化模板体',
 '面试什么话不能说': '禁忌清单体/雷区警示体',
 '实习谈薪': '流程攻略体/话术清单体',
 '涨薪案例': '数字对比体（3w→9w→18w）/身份标签体',
 '推迟面试话术': '避雷体/教科书模板体',
 '会议礼仪发言': '万能公式体/礼仪清单体',
 '面试改期': '避雷体',
}

# ================= Sheet1: 内容表 =================
ws1 = wb.active
ws1.title = '内容表'
h1 = ['关键词', '来源', '新增/已收', '标题', '作者/站点', '热度', '核心要点', '标题套路', '链接']
ws1.append(h1)
rows_sorted = sorted(rows_out, key=lambda r: (r[0], r[2] != '今日新增'))
for r in rows_sorted:
    kw, src, flag, title, author, hot, url = r
    point = POINT_HINTS.get(kw, '')
    pattern = TITLE_PATTERN_HINTS.get(kw, '')
    ws1.append([kw, src, flag, title, author, hot, point, pattern, url])
style_header(ws1, len(h1))
widths1 = [14, 8, 10, 34, 16, 8, 22, 24, 46]
for i, wd in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = wd
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
    if row[2].value == '今日新增':
        for cell in row:
            cell.fill = NEW_FILL

# ================= Sheet2: 根词表现表 =================
ws2 = wb.active
ws2 = wb.create_sheet('根词表现表')
h2 = ['根词', '类型(本轮身份)', '新增条数', '总抓取条数', '命中率', '平均热度(估)', '升级/退休', '备注']
ws2.append(h2)
ROLE = {
 '职场表达': '种子', '面试技巧': '种子', '面试什么话该说': '种子', '面试什么话不能说': '种子',
 '实习谈薪': '长空档活跃复查', '涨薪案例': '长空档活跃复查',
 '推迟面试话术': '候选首投', '会议礼仪发言': '候选首投', '面试改期': '既有活跃(顺带命中)',
}
UPGRADE = {'推迟面试话术': '候选→活跃(66.7%)', '会议礼仪发言': '候选→活跃(77.8%)'}
NOTE2 = {
 '职场表达': '种子词同日多轮衰减符合历史规律',
 '面试技巧': '0新增,种子词严重饱和,建议下次拉长复查间隔',
 '面试什么话该说': '仅1条新增,趋于饱和',
 '面试什么话不能说': '3条新增,含2条负面判读类新内容',
 '实习谈薪': '27.3%为该词历史最低值(此前均>=40%),首次<40%,维持活跃挂观察',
 '涨薪案例': '57.1%,健康水平,加拿大/德国海外账号供给较多需留意相关性',
 '推迟面试话术': '66.7%远超阈值,升级为活跃;命中率口径为XHS+web合计',
 '会议礼仪发言': '77.8%全场最高,但9条中约4条为外企英语内容,相关性需人工复核',
}
AVG2 = {'职场表达': '约2.3万', '面试技巧': '约6千', '面试什么话该说': '约6千', '面试什么话不能说': '约6千',
        '实习谈薪': '约900', '涨薪案例': '约500', '推迟面试话术': '约350', '会议礼仪发言': '约2600',
        '面试改期': '—'}
for kw in ['职场表达', '面试技巧', '面试什么话该说', '面试什么话不能说', '实习谈薪', '涨薪案例', '推迟面试话术', '会议礼仪发言', '面试改期']:
    if kw not in stat:
        continue
    v = stat[kw]
    tot = v['xhs_tot'] + v['web_tot']; new = v['xhs_new'] + v['web_new']
    pct = f"{round(new / tot * 100, 1)}%" if tot else '—'
    ws2.append([kw, ROLE.get(kw, '—'), new, tot, pct, AVG2.get(kw, '—'), UPGRADE.get(kw, '—'), NOTE2.get(kw, '')])
style_header(ws2, len(h2))
widths2 = [16, 16, 10, 12, 10, 12, 20, 40]
for i, wd in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = wd

# ================= Sheet3: 词库验证表 =================
ws3 = wb.create_sheet('词库验证表')
h3 = ['长句', '场景类型', 'a-下拉/大家都在搜命中', 'b-结果页数量级', 'c-前排10篇赞藏(<500达标)', '判定结果', '意图强度', '竞争密度', '备注/答案空缺']
ws3.append(h3)
VERIFY_ROWS = [
 ['面试官说3天内给回复有没有希望', '事件',
  '命中同级变体:面试官说三天内给回复有戏吗/面试完说3天内给回复/面试结束后说三天内给答复/面试官说3个工作日给回复', '20条', '前排多数<500,仅2篇>500(3506/3131,且偏泛化)', '已验证', '高', '低',
  '前排出现"HR说三天给答复第几天没信就凉"(384赞)首条天数拆解苗头,但仍是判读层,未给动作模板,与跨5轮空缺同源'],
 ['口头offer后怎么催进度', '事件',
  '命中完全同字面原词笔记(昕哥说就业连发2条同标题)+催书面offer话术/发了口头offer怎么催/国企口头offer后怎么催进度', '20条', '多数<500,3/13超500(949/734/1652)', '已验证', '高', '低-中',
  '★重要突破:精确到"口头offer后"处境的问法已形成独立搜索习惯,前排逐字可复制话术仍稀少,建议优先出稿填补跨4轮的催进度空缺'],
 ['hr暗示你已经被录用了', '事件',
  '命中同级变体但含1条偏离题(hr问你为什么放弃offer)', '20条', '不达标,前排多篇>500(1342/3465/3183/2970)', '放弃', '—', '—',
  '判读型"HR暗示信号"子赛道已高度饱和,与08-16/08-17"判读型内容止步于识别"结论一致'],
 ['微信怎么问面试结果话术', '事件',
  '命中同级变体:怎么礼貌的问面试结果话术/加了微信怎么问面试结果/面试主动问结果话术/微信询问面试结果怎么说合适', '20条', '不达标,前排多篇>500(1208/1146/3506/5622)', '放弃', '—', '—',
  '"要不要问/怎么问结果"方法层已饱和,但对比已验证词"口头offer后催进度"说明挂载具体处境/节点的问法仍有空间(方法论发现)'],
]
for row in VERIFY_ROWS:
    ws3.append(row)

ws3.append([])
ws3.append(['本轮收割新词清单(关键词池-候选, 8个)'])
for w in ['催offer话术', '涨薪成功案例分享', '晋升调薪', '开会发言礼仪', 'hr录用信号', '面试结果询问话术', '会议发言万能公式', '面试话术秘招']:
    ws3.append([w])
ws3.append([])
ws3.append(['本轮收割新词清单(词库-候选长句, 11个)'])
for w in ['hr说明天发offer是不是稳了', '国企口头offer后怎么催进度', '怎样催offer比较礼貌', '面试完傻等通知怎么破',
          'hr问你为什么放弃offer怎么回答', '面试延迟再约怎么说', '已经答应面试时间又要改期怎么说',
          '会议礼仪的稿子怎么写', '面试完一周没消息还要不要投别的', '实习期间可以主动提涨薪吗', '涨薪多少合适']:
    ws3.append([w])
ws3.append([])
ws3.append(['本轮最大答案空缺'])
ws3.append(['面试完等通知期间怎么开口催进度的可复制原话——跨5轮验证仍为空白，但本轮找到潜在突破口"口头offer后怎么催进度"(已验证,已收录),建议下轮围绕该突破口深挖具体天数节点话术'])
ws3.append(['方法论发现'])
ws3.append(['泛化的"怎么问结果"类母题已饱和,但挂载具体处境(口头offer后/3天节点后)的精确问法仍有独立搜索需求且竞争更低——搜索词精确度与竞争密度成反比的新证据'])
ws3.append(['运维突破'])
ws3.append(['笔记详情页/评论区连续5轮不可达问题已解决：改用搜索结果页read_page输出中自带xsec_token的/search_result/{id}链接跳转,而非手工拼接裸/explore/{id},详情页与评论区均可正常加载(本轮实测3714条评论正常展开)'])

style_header(ws3, len(h3))
widths3 = [30, 10, 34, 10, 26, 10, 8, 8, 40]
for i, wd in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = wd
for row in ws3.iter_rows(min_row=1, max_row=6):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

out_path = D + '职场表达与面试技巧_2026-08-18_run2.xlsx'
wb.save(out_path)
print('saved:', out_path)
