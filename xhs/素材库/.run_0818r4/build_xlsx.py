# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D = '/sessions/ecstatic-festive-cori/mnt/素材库/'

wb = Workbook()
HEAD_FILL = PatternFill(start_color='FFDCE6F1', end_color='FFDCE6F1', fill_type='solid')
HEAD_FONT = Font(bold=True)
NEW_FILL = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'

ws1 = wb.active
ws1.title = '内容表'
h1 = ['关键词', '来源', '新增/已收', '标题', '作者/站点', '热度', '核心要点', '标题套路', '链接']
ws1.append(h1)

POINT_HINTS = {
 '职场表达': '汇报进度高频短语/沟通话术',
 '面试技巧': '综合面试建议',
 '面试什么话该说': '经典面试问题标准答案清单',
 '面试什么话不能说': '面试雷区禁忌清单',
 '加薪话术': '加薪谈判话术/向上沟通模板',
 '结构化汇报': '结构化表达框架/汇报方法论',
 '会议发言万能公式': '临场发言公式/上台发言技巧',
 'hr录用信号': '录用信号判读清单',
}
TITLE_PATTERN_HINTS = {
 '职场表达': '数字短语清单体',
 '面试技巧': '清单体万能话术',
 '面试什么话该说': '清单体标准答案',
 '面试什么话不能说': '避雷警示体',
 '加薪话术': '话术清单体/情境对话体',
 '结构化汇报': '"工作能力强的人都..."母题反复复用/方法论清单体',
 '会议发言万能公式': '万能公式体',
 'hr录用信号': '信号清单体/征兆判读体',
}

def link(nid, token):
    return "https://www.xiaohongshu.com/search_result/" + nid + "?xsec_token=" + token + "&xsec_source="

rows = [
 ("会议发言万能公式","职场会议控场500句实用句式","69229e96000000001e006a8e","ABhkoJ5nHSsKcyAexU3bJp_24e6I3DR15J6xq5aqzpCKQ=","Dimoo","—"),
 ("会议发言万能公式","线上会议不再尴尬！100句帮你高效沟通！","67125e950000000021009451","AB5sAK52K7Au3KG5MeshDb3ly_7kb4XVxjqg0edoPiZ08=","小Mia_","1535"),
 ("会议发言万能公式","任何场合都不慌的万能发言公式","6a54e5880000000008027740","ABvH1EJVQVpMIscVIyksCjmyK-lkZ_xerQYcjKvtzH1-s=","桃子练表达","280"),
 ("会议发言万能公式","工作能力强的人全部擅长上台发言","6a22d2cd000000000702a08c","AB2-zK3TMjQ9weQqJFH3vJ8-jTpwJk21jhnmEU2diYnvg=","表达力精进HST","—"),
 ("会议发言万能公式","开会即兴临时发言万能模版","6908219800000000040100af","ABEzCq-alzmEinhN5U051O34W_t2S2Cpx9nnjGY_-hW2k=","绝知笔记","579"),
 ("会议发言万能公式","上台发言4个万能公式","694c9125000000001e011974","ABQTjiGA6A26rXXTB_tr1H3H_V9yHCEnNsyODf7PYnKgY=","商业思维-田野","—"),
 ("职场表达","6个让老板秒懂进度的工作汇报高频短语","6a71b8c2000000002402d72a","ABJ67AJwpaEbbSDyndwOBnXkD3iOf5oez-8phBwLe9hZY=","Rainbow聊外企","2348"),
 ("面试什么话不能说","面试别太老实！HR淘汰人，从来不只看能力","6a4ee8830000000011012b17","ABRx_IccmX9kdD97OzyHFV2M2j8s87QJGCJHf8IIBUpzQ=","新建练口才","454"),
 ("面试什么话该说","10大经典面试问题，HR想听什么","68c1198a000000001b033e7a","AB7axxOzgLY1W4r_XU4ANoBQRMZBjofHo7kLAtIlnyapA=","职场教练郭钦","1.7万"),
 ("加薪话术","100% 真实外企加薪沟通版（自然、可复用）","697b77b6000000000c035f8d","ABk7u19hQjmrKPc6PNeLpEGo4g8c_q3N3jKKpN0oXxkfo=","果果Ella","374"),
 ("加薪话术","让老板极度舒适的加薪话术 亲测有效","6915b230000000000d03a72a","ABI6CzYzOiHYMQR5ZX0_hZPYDFpO_XzcCMWzDCvfLJ4y4=","胡诌竹子","343"),
 ("加薪话术","提加薪必看！3个让领导无法拒绝的硬核话术","68a6ccfd000000001d018ead","ABzMt8MyHEI6SrozO9xhZuqVpT4xmSpVjyIZRV_JSfMAk=","职场外援老肖","116"),
 ("加薪话术","跳槽谈薪掌握这些技巧，薪资直接涨40%","69af7e26000000000e03c805","AB_Ikwr_oU4Dcge5T9H1i0wfeW1fi0IIEJQpS68jiEFZU=","橙子的职场笔记","224"),
 ("加薪话术","向上管理｜如何优雅地和leader谈涨薪","67ac3691000000002901db40","AB6tLLM5fPFZieWIus-VDCqmk6UnyHoqO6N-sEMkjE6p4=","野生澄（重启人生版）","820"),
 ("加薪话术","如何谈薪才能做到既不亏又不黄？","69b7be97000000001f001b5a","AB-_QWW-qMGrRIhO5HhjzFqf5BwnXlqXHZgQZFvzVyoW4=","夏鹏老师","770"),
 ("加薪话术","涨薪指南这样跟领导谈加薪 万无一失","69926b40000000001a0265b7","ABRyigYanIGxIlNKtMEdPcNDvneFwgvxh_QZ4x9pYSdq8=","有冰姐在","5302"),
 ("结构化汇报","PPT汇报万能的6大逻辑结构化表达","69b9f5420000000022026d7a","ABxCakbYhyYN5sw_5NAzXFJOxeUD2wA2YzrUFuTbRjYOM=","加薪PPT","—"),
 ("结构化汇报","大厂结构化汇报，leader一听就知道你是高潜","6a3ba8fd000000000f01f94c","ABpEawNYaVi9TG4Kvzy0HHuDemDeksu4dWQdw_77vsC_s=","思思姐姐-晋升述职教练","—"),
 ("结构化汇报","如何提升结构化表达？3步成为表达高手！","69f327e70000000036003f75","ABvjXng2EjCiS1emry2jzk88DZPupsWCQQ8Fx4P1ZTnkA=","一支粉笔讲管理","9157"),
 ("结构化汇报","向上汇报的结构化思维","69bc00f800000000220243d5","ABGc6s9hoZ_Jl_3-7aUB_fytaY0gXOVZXauZWlBRqOO8Y=","九柏-破局职场瓶颈","—"),
 ("结构化汇报","结构化汇报-职场人升级打怪的制胜宝典","661a8730000000001a01427c","ABME5gAzgZrKA1JxffbbobcdcxcDyb2tbEEWs-xbxMs8g=","木夕游墨","—"),
 ("结构化汇报","设计师：结构化表达，太香了！","694bd23b000000001e00c06d","ABe8MEio87HZrcFm1gXR2mzbYFuxeg2b6zoFJwO6QNbX0=","叙茶拾光（设计照进生活）","—"),
 ("结构化汇报","如何完成一次高质量汇报分享？","67c6932c000000000d0160e6","ABsoTlOE85Zqfl7OyKAWBh02P4WXmfh_W9sycCPy8SfLI=","宫了了的思维工具箱","336"),
 ("结构化汇报","高手的工作汇报，都是用结构化表达包装的！","6659457d00000000050056dd","ABr4VKqUBtCILnx6inUPSLvKNeJCSYw1eZngrLhUR_azA=","职场一只小桔子","781"),
 ("结构化汇报","无论汇报、述职还是面试，记住这5步法","6863e55800000000170341fd","ABMuv1d96Foz47oIyMWl2QfGC-c1NVGtFjojXg4qwkgtw=","Elsy职场力","300"),
 ("结构化汇报","工作能力强的人都擅长结构化汇报","69ca64940000000021013b52","AB3wtcZd1-WRHAqhBNND9KeddK9BwJpqN6B6rydAUqVmE=","培训杰森老师","—"),
 ("结构化汇报","工作能力强的，全都擅长结构化汇报","66ebef74000000000c018908","AB2qBDm6dPTcZyq8GGunpKwLOeGoaI-c5593hD7VLxkrc=","麻豆豆吖","1460"),
 ("结构化汇报","麦肯锡：Storyline让核心观点一目了然","6804c451000000001c01ddc8","ABAMDPOW4mSyAq7BNqXhmo32lFuEzHtd7Cp1bkCrZ3PRc=","麦家PPT","436"),
 ("结构化汇报","PPT恐惧症进-最加分的汇报结构让领导秒懂","69f8851a0000000037036677","ABuw_pgE8NFaCL2ge9TDK_dBnuM4aRjIhrNiI5k6tBPiA=","长缨不嘤嘤","7512"),
 ("结构化汇报","怎么让领导知道你干了很多活儿：结构化汇报","6a81eeac000000003301a1a4","ABEdG0crDJWcc91KSvYHk15I6ZH_GHzdlbNBySaVaq__M=","Molly酱破壳中","—"),
 ("结构化汇报","中层汇报没故事线=白干！","6a058ebb000000003701cbb6","ABX98OOVjvrVL3rZGpK4x_qH-CZGGwlq-wbQK4DqutnFI=","Nina管理笔记","307"),
 ("hr录用信号","真正想招你的公司会非常明显！","69676d6b000000001a02c8c6","ABbDLQPS_2V4r_lvZ5myGTyr2pGGG0CFOF4cowi9uikEE=","(未知)","576"),
 ("hr录用信号","恭喜你，被录取啦","6a05b1fa00000000060202a9","ABX98OOVjvrVL3rZGpK4x_qB3-TGGklt-CKWiTverUJY8=","予你offer_研究所","—"),
 ("hr录用信号","有以下情况，证明你要被录用了","69cba4db000000001d01d75b","ABbgTHHddrmRNIelr_tl_YfZ1wKFl1jlMMDqSacdl630Y=","一只小职聊","—"),
 ("hr录用信号","这7个信号，是要给你发offer的征兆","6a221df200000000220259ba","AB2-zK3TMjQ9weQqJFH3vJ86GGTjBabcEJBZ3V0c7zgnk=","欧Sir职场进化","—"),
 ("hr录用信号","北美面试官如果真的想招你，会有这6个信号(海外场景)","696a848d000000000b00a87f","ABMI9PdbRayK7cSMy6_X8jbSvkK5P72zbUDKdFsIqC_8I=","Auroraaa呀","—"),
 ("hr录用信号","HR说出这些话，说明你的offer稳了","6840fc5c0000000022027790","AB8gyJRZg2QSVeM_Ce60sRb-V7dM0hZ85l591p6gBwAug=","人力牛马","—"),
 ("hr录用信号","北美HR判断可培养候选人的隐形信号！(海外场景)","69809a9a000000001a03145e","ABkwVhSwSuEceyASCD9S0NgPmgXgx5LHjM-VA4o3NRduo=","北美HR-Zoe捞人啦","—"),
 ("hr录用信号","offer审批的3个秘密，HR不会告诉你","69e0b4fc000000001f0001b4","ABI3bPvVoN11x29eMZMUxszyQjYqyFHwkUVSdqWthpuvU=","夏花谈职场","—"),
 ("hr录用信号","加拿大银行面试结束语解密(海外场景)","68c11070000000001d0258e1","AB7axxOzgLY1W4r_XU4ANoBd1YtzMhrzi9oLeljqwaWAA=","土豆国国王","—"),
]

kw_order = {"职场表达":0,"面试技巧":1,"面试什么话该说":2,"面试什么话不能说":3,"加薪话术":4,"结构化汇报":5,"会议发言万能公式":6,"hr录用信号":7}
rows_sorted = sorted(rows, key=lambda r: kw_order.get(r[0], 99))
for kw, title, nid, token, author, hot in rows_sorted:
    ws1.append([kw, '小红书', '今日新增', title, author, hot, POINT_HINTS.get(kw,''), TITLE_PATTERN_HINTS.get(kw,''), link(nid, token)])

style_header(ws1, len(h1))
widths1 = [16, 8, 10, 36, 18, 8, 22, 26, 46]
for i, wd in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = wd
for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
    if row[2].value == '今日新增':
        for cell in row:
            cell.fill = NEW_FILL

ws2 = wb.create_sheet('根词表现表')
h2 = ['根词', '类型(本轮身份)', '新增条数', '总抓取条数(估)', '命中率', '平均热度(估)', '升级/退休', '备注']
ws2.append(h2)
ROLE2 = {
 '职场表达': '种子(今日第4轮)', '面试技巧': '种子(今日第4轮)', '面试什么话该说': '种子(今日第4轮)', '面试什么话不能说': '种子(今日第4轮)',
 '加薪话术': '长空档活跃复查(16天)', '结构化汇报': '长空档活跃复查(16天)',
 '会议发言万能公式': '候选首投', 'hr录用信号': '候选首投',
}
STAT2 = {
 '职场表达': (1, 20, '5.0%'), '面试技巧': (0, 20, '0.0%'),
 '面试什么话该说': (1, 20, '5.0%'), '面试什么话不能说': (1, 20, '5.0%'),
 '加薪话术': (7, 20, '35.0%'), '结构化汇报': (15, 20, '75.0%'),
 '会议发言万能公式': (6, 20, '30.0%'), 'hr录用信号': (9, 20, '45.0%'),
}
AVG2 = {'职场表达': '约2.3万(存量均值)', '面试技巧': '约6千(存量均值)', '面试什么话该说': '约6千(存量均值)', '面试什么话不能说': '约6千(存量均值)',
        '加薪话术': '约1178(7条样本)', '结构化汇报': '约2536(8条有热度样本)', '会议发言万能公式': '约798(3条样本)', 'hr录用信号': '约576(1条样本)'}
UPGRADE2 = {'hr录用信号': '候选→活跃(45.0%超40%阈值)'}
NOTE2 = {
 '职场表达': '今日第4次运行,严重饱和,与前3轮(0/5/3新增)一致衰减',
 '面试技巧': '0新增,今日第4次运行,已连续多轮归零',
 '面试什么话该说': '1条新增,趋于饱和',
 '面试什么话不能说': '1条新增,趋于饱和',
 '加薪话术': '35.0%略低于历史82.4%但仍产出7条含万级赞样本(涨薪指南5302),维持活跃',
 '结构化汇报': '75.0%,再次验证长空档活跃词优于同日重复种子(种子今日0-5% vs 本词75%),15条新增中"工作能力强的人都擅长结构化汇报"母题反复出现,内容供给同质化明显',
 '会议发言万能公式': '30.0%未达40%阈值,候选首投未升级;且与既有活跃根词"即兴发言"内容高度重叠,边际增量存疑',
 'hr录用信号': '45.0%超阈值升级为活跃;9条新增中3条为北美/加拿大海外场景,相关性需持续复核',
}
for kw in ['职场表达','面试技巧','面试什么话该说','面试什么话不能说','加薪话术','结构化汇报','会议发言万能公式','hr录用信号']:
    new, tot, pct = STAT2[kw]
    ws2.append([kw, ROLE2[kw], new, tot, pct, AVG2[kw], UPGRADE2.get(kw, '—'), NOTE2[kw]])
style_header(ws2, len(h2))
widths2 = [18, 20, 10, 14, 10, 20, 24, 46]
for i, wd in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = wd
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

ws3 = wb.create_sheet('词库验证表')
h3 = ['长句', '场景类型', 'a-下拉/大家都在搜命中', 'b-结果页数量级', 'c-前排10-12篇赞藏(<500达标)', '判定结果', '意图强度', '竞争密度', '备注/答案空缺']
ws3.append(h3)
VERIFY_ROWS = [
 ['国企口头offer后怎么催进度', '事件',
  '命中同级具体变体:国企催offer话术/国企社招口头offer/国企offer审批期间可以催吗/没收到offer可以催吗', '约15条相关但多为泛化催offer内容,国企专属供给稀薄(少量<50)', '前排12篇中4篇>500(1652/3506/2527,另1篇2302属面试技巧串词非本题)其余<500', '已验证', '高', '中',
  '国企具体限定词的专属供给稀薄但泛化催offer母题已有占位,竞争密度判定为中(非低);详情页抽检1篇("手头有其他offer后如何礼貌催offer分享")评论区仅1条"学到了!"未暴露新处境'],
 ['怎样催offer比较礼貌', '事件',
  '命中同级变体:可以用一个offer去催另一个吗/如何委婉催offer/催offer流程话术/怎么跟hr催offer', '"催offer"泛化母题内容密集,约12+篇直接相关,接近饱和', '前排12篇中5篇>500(1652/1033/734/960/707),普遍<500不成立', '放弃', '—', '—',
  '与08-18run2"微信怎么问面试结果话术"放弃案例同构,第2次验证"泛化催offer/问结果母题已饱和,但挂载具体处境(国企/口头offer后/3天节点)的精确问法竞争更低"的方法论'],
 ['hr问你为什么放弃offer怎么回答', '事件',
  '大家都在搜首条即命中几乎逐字同词"hr问放弃offer原因怎么回答"', '结果页多为"婉拒offer/拒绝offer"邮件模板等泛化母题,聚焦"被追问原因怎么回答"的专属内容仅1-2条', '前排出现多篇泛化拒绝offer赛道爆款:嘴笨的人拒绝offer反而更顺利1.8万/收到offer后怎么高情商回复hr 5319/婉拒offer后HR来信2082/婉转拒绝offer附话术1807/国人总爱过度解释离职原因6681', '放弃', '—', '—',
  '★与08-17run4"面试官问自我评价"证伪案例同构(第2例):条件a强命中(甚至近乎逐字)不能替代b/c验证,推断出的"更像搜索词"的表述仍必须过完整三条件'],
]
for row in VERIFY_ROWS:
    ws3.append(row)

ws3.append([])
ws3.append(['本轮收割新词清单(关键词池-候选, 6个)'])
for w in ['结构化汇报模板', '结构化汇报公式', 'hr积极信号', 'hr录取信号', '会议发言思路', '工作会议发言']:
    ws3.append([w])
ws3.append([])
ws3.append(['本轮收割新词清单(词库-候选长句, 10个)'])
for w in ['国企offer审批期间可以催吗', '国企口头offer多久发正式通知', '没收到offer可以催吗', 'hr一直不回复offer结果',
          'hr口头offer没后续怎么问(★与跨轮最大空缺高度相关,建议下轮优先验证)', '明确回复hr offer了不去可以吗',
          '结构化汇报怎么写', '给员工加薪话术', '申请调薪的简短话术', '涨薪申请理由怎么写']:
    ws3.append([w])
ws3.append([])
ws3.append(['本轮最大答案空缺'])
ws3.append(['面试完等通知期间怎么开口催进度的可复制原话——跨6轮验证仍为主空白;本轮已验证"国企口头offer后怎么催进度"(竞争密度中,非纯清白地带),并新收割"hr口头offer没后续怎么问"作为下一个待验证的潜在突破口'])
ws3.append(['方法论发现(本轮强化)'])
ws3.append(['条件a"命中"只是入场券,真正决定竞争密度的是内容供给端是否已把该具体处境写滥;条件a强命中甚至逐字命中都可能是假信号(第2个同类证伪案例);同一天内种子词边际回报衰减速度远快于长空档活跃词(本轮同日对比:结构化汇报75% vs 四个种子词今日第4轮0-5%)'])
ws3.append(['运维记录'])
ws3.append(['评论区原话抓取本轮验证性尝试2篇(延续08-18run2的xsec_token跳转方法):1篇详情页评论区显示"这是一片荒地"未渲染出可读评论;另1篇仅1条评论未暴露新处境;评论区原话.csv本轮无新增,方法论本身仍成立,受限时间预算未扩大抓取范围'])

style_header(ws3, len(h3))
widths3 = [32, 10, 34, 24, 30, 10, 8, 8, 44]
for i, wd in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = wd
for row in ws3.iter_rows(min_row=1, max_row=len(VERIFY_ROWS)+1):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

out_path = D + '职场表达与面试技巧_2026-08-18_run4.xlsx'
wb.save(out_path)
print('saved:', out_path)
