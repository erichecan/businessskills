# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")

def style_header(ws, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============ Sheet 1: 内容表 ============
ws1 = wb.active
ws1.title = "内容表"
headers1 = ["关键词","来源","新增/已收","标题","作者/站点","热度","核心要点","标题套路","链接"]
ws1.append(headers1)

def url(nid, token):
    return f"https://www.xiaohongshu.com/search_result/{nid}?xsec_token={token}&xsec_source="

content_rows = [
("面试反杀面试官","小红书","新增","压力面试，是你甩开对手的好机会（第三期）","会说话的豆腐","4.4万","把压力面试重新定义为筛选弱者的机会，正向心态转化","反差判断+系列号",url("69a11684000000001a026624","ABZ-ut7Qx7gEc2431nczyoanX_l0BWznSZdUP5dzRT0RI=")),
("面试反杀面试官","小红书","新增","用 AI 面试作弊，还拿到了亚马逊的offer","晓白和林亦","8307","AI辅助面试实操经历分享","反差禁忌词+结果导向",url("67c81863000000002a00fe39","ABVRlWIBdBAO3f4eV6TNqA65as5H8iY3nP1b2kbzR3CX8=")),
("面试反杀面试官","小红书","新增","怎么让面试官对你增加好感","Mr Jonathan","7215","面试官心理博弈技巧，建立好感的具体动作","疑问句+承诺",url("69af6c3d000000001503b0dc","AB_Ikwr_oU4Dcge5T9H1i0wSbaycebsNTwTcEHRDeJhXQ=")),
("面试反杀面试官","小红书","新增","一个小技巧 面试官也想帮你拿offer","Mr Jonathan Office","533","单点技巧+利他视角反转","数字+反转视角",url("6a46e2b4000000000e021800","ABpOtIMLDH4jJTZZrsOXw4pSxtNZjHgsPQD6_NLOi8Wi0=")),
("面试反杀面试官","小红书","新增","假如面试官说真话","Mr Jonathan","5893","面试官视角揭秘潜台词","假设句+揭秘",url("69ab5d38000000001d013859","ABnKti2O_OHU3FClrBkxVLeyE1h7t1y_bprhK7d7uPPUU=")),
("面试反杀面试官","小红书","新增","拒绝题海战术 把面试官框在你的思路里","Mr Jonathan","2741","反套路刷题，主动引导话题框架","反直觉+动词强行动",url("695346160000000021028055","ABNrycHnQDFJUd8jY2-YmsIk0Gwh5a3pZVwUGKhQwirjQ=")),
("面试反杀面试官","小红书","新增","把面试官绕在你的思路里","Mr Jonathan","3702","主导对话节奏的话术框架","动词强行动",url("69b20c06000000001503b876","ABdb0EnkF1RuquHphATtQx0wQUWvh7i0SIbt4OD8VJ8T8=")),
("面试反杀面试官","小红书","新增","让面试官知道咱们中国人软硬实力兼备","Mr Jonathan","2617","身份认同+文化自信角度","身份共鸣",url("6967e0c400000000220388d8","ABbDLQPS_2V4r_lvZ5myGTyr7ImdlUIHY50M_0J6RsP8k=")),
("面试反杀面试官","小红书","新增","把压力给到面试官！","iPig","2365","攻守转换的心态调整","反转+感叹号",url("6909ed5c0000000003038a80","ABHD4wwLZpEZgDk0Wt_qHNu6VOItXnMY3rfF807ZkulKY=")),
("面试反杀面试官","小红书","新增","高压面试如何用“停顿”反杀面试官？","华裔高管Demi林-觅得林","23","具体技巧(停顿)+身份背书(高管)","权威身份+具体技巧",url("69fb95ff000000000f03ac00","ABySzai4nLI67tf-ONK9YhYwvxmTDBNbwLzuZpwtLzUDc=")),

("职场表达","小红书","新增","做了个训练讲话的网页-2.0版本-第2期","陈红叶","2690","自制工具型内容，可复制使用","工具体+系列号",url("6a6ddae6000000002c005de5","ABk31yDVGRaeJy7muaLFSpuKHReolTCeosjRf0BZbWHTA=")),
("职场表达","小红书","新增","麦肯锡顾问：废话太多怎么办？（连载01）","麦肯锡朱老丝","689","权威咨询背景+具体症状","权威身份+症状疑问",url("69b7e24200000000230203fd","AB-_QWW-qMGrRIhO5HhjzFqYwl0oXkMGUbOA2qRb2NXHg=")),
("职场表达","小红书","新增","精选外刊丨人一定要大量频繁的说话","每日精选外刊","2290","强调多说多练的表达训练观点","观点断言体",url("6a4bb169000000001603c6ad","AB78zpW53s2LixtvzaoSp1XjMYU_ml5MB80jklPiPD4sQ=")),
("职场表达","小红书","新增","那些外国同事每天都用的地道表达（十八）","Joan要摸鱼学英语","919","英文职场表达系列合集","系列号+身份差异",url("6a460e3d00000000170080f5","ABpOtIMLDH4jJTZZrsOXw4pd9rk9jmtwZgZy_JtsMy19c=")),
("职场表达","小红书","新增","白女同事亲授如何在会议里“自然抢到话语权”","Robin漂流记","381","跨文化会议话语权技巧","身份标签+具体场景",url("694938ee000000001e02f9a8","ABBGSvQuwpmbnGZmzU7E7tX77QnTFkDt9fqnk-1YZ1T4I=")),
("职场表达","小红书","新增","表达能力，决定了你的职场上限","工作笔记本","7.1万","高赞爆款，表达能力与职场天花板挂钩","断言+利益直给(上限)",url("69db808200000000230202d1","ABU_FwWTZVdUj5at-q8TnB81_C1sWF7DlMBFywCwD9IZM=")),

("面试技巧","小红书","新增","1个面试不卡壳的万能话术","Mr Jonathan Office","659","单一万能话术，降低使用门槛","数字1+万能话术",url("6a6bc91a0000000008009c00","ABd2kIn4bXJvoibTlK7X9tR06ZjVOZsnndGVnyPAdLeHA=")),

("面试什么话不能说","小红书","新增","面试时千万不要说这句话","Dr Jeff Bogaczyk","193","单句禁忌提醒，制造紧迫感","禁忌警告体",url("69ab30d80000000026033031","ABnKti2O_OHU3FClrBkxVLe2Pb7IO-GmXnOmdaBMcrh24=")),
("面试什么话不能说","小红书","新增","面试时最让HR讨厌的3种自我介绍","职得谈","133","HR视角的雷区归纳","数字+HR视角信息差",url("6a031219000000003701faf1","ABZeUe5sJ0ci7e_0MDNUPghWNFXp2uIY2_7opemJhVQMQ=")),
("面试什么话不能说","小红书","新增","面试像聊天 反而拿offer","key319","1.2万","反直觉：放松聊天式面试更易成功","反直觉断言",url("67188ff80000000021002cbf","ABxshC36zcitx4_nQnge_7lAXovNcidBmox--OuA-16j4=")),

("面试被hr阴阳怪气","小红书","新增","面试感到不被尊重，那是贵人面试官叫你快跑","香辣虾饲养员","811","把不被尊重的面试重新定义为避坑信号","反转判断+身份标签(贵人)",url("680a0d14000000000b01cb71","ABOarNiFIToxm-YZiHQ7vT5tQyADid8hBONduqOyXhoBQ=")),
("面试被hr阴阳怪气","小红书","新增","我开始在面试中怼面试官了","我命油天天","255","第一人称经历，主动反击叙事","第一人称行动体",url("69d5521c0000000022003438","ABV2YhTVTfSTK65SR4NFFjKI4e1eZoUlefh_J_wwe3wV4=")),
("面试被hr阴阳怪气","小红书","新增","面试被问不稳定，怎么回答","职场研究小Go","—","具体刁钻问题的应对方法","疑问句+具体场景",url("69368a43000000000d00c8d2","ABKj4CHkEKquroeemQLuoiAwcfwIvPSwiCFiIgdpSIsM4=")),
("面试被hr阴阳怪气","小红书","新增","遇到无礼的面试官请果断礼貌的回怼过去！","Ken的趣识堂","—","礼貌回怼行动指令","祈使句+感叹号",url("69dcd3f40000000023015ca6","ABlXCt1FSztGRKYJ7UtCpnYBxUr_PxaTkaTT_AUjQsYo4=")),
("面试被hr阴阳怪气","小红书","新增","为什么有些HR面试喜欢习惯性打压？请看vcr：","一只小羊","—","现象归因+证据体(vcr)","疑问句+证据体",url("6a1a90ef000000003503aa04","ABNYw0WGPP7PwQbkwmzu0-MvsQHw3mfCuSAMXsSZIqZf0=")),
("面试被hr阴阳怪气","小红书","新增","如何礼貌阴阳男hr，留子和国内hr沟通被创飞","北美打工momo（old grad）","—","跨文化沟通冲突的第一人称吐槽","第一人称+身份标签(留子)",url("68006d90000000001e0060e4","AB52wtiI_BSQNzpsbLbNL-qw0A80yz9m55XGQi3tgCmdo=")),
("面试被hr阴阳怪气","小红书","新增","救命！面试总踩坑，原来是没懂 HR 这些话背","北漂运营打工仔","1.5万","HR潜台词解码，避坑指南","惊呼体+潜台词揭秘",url("68518792000000000c03aa2c","ABhkGh6wXMM_UuUN3Nd5e0F8mFEeH2DPhtMC6Kn_Qao4Q=")),
("面试被hr阴阳怪气","小红书","新增","投简历时被HR阴阳到笑不出来","面面面","—","第一人称吐槽体，情绪共鸣","第一人称情绪体",url("680df4560000000022027fe8","ABIfg8SJxleEDJq-5xulGz_ONDWlefpJVWLo434BetnpQ=")),
("面试被hr阴阳怪气","小红书","新增","🆘面试之中，遇到这些一定要主动叫停！","建林聊职场","1977","警示+行动指令","emoji警示+祈使句",url("69fa0fd2000000003601d892","AB1NLaeWrchVTAuM6tVPcgxVPQLaPwaYwVJe96lMrHtrE=")),
("面试被hr阴阳怪气","小红书","新增","5句话，反杀阴阳你的人","摆渡人L","165","具体逐句话术清单","数字+可抄话术",url("6a0d85b20000000006020180","ABPcuE_PJ-3g8HlVeSNJWQrDmHRrjw3gTT3DOfumhcyP8=")),
("面试被hr阴阳怪气","小红书","新增","🇨🇦刚刚结束了一场让我无地自容的面试","葱丝姜丝萝卜丝","152","第一人称糟糕经历叙述","第一人称情绪体+地域标签",url("697a5f1d000000000e00f841","ABlzAmUXndJafN8esrtyIiU3CouR57-45uPklIwnVxffM=")),
("面试被hr阴阳怪气","小红书","新增","有些女面试官的微妙恶意","又见夏天","178","隐晦恶意的观察归纳","观点断言体",url("69ce0c450000000022028ce5","ABWirANUTJFhvHvOjSkh3Rgf274XYflldE4h9vvj0dn4k=")),

("求职面试高频题库","小红书","新增(候选)","FP&A面试高频题｜面试前一定要过一遍","毛小毛","—","垂类岗位(FP&A)高频题整理","垂类身份+行动指令",url("6a78b194000000002500e03b","ABgC4AG4uV8pNGdRUoacucZqIzY30nvrjWRzB0nYndBVc=")),
("求职面试高频题库","小红书","新增(候选)","AI Engineer面试建议这样准备","独坐芋泥山","—","技术岗面试准备指南","垂类岗位+建议体",url("6a27d8df000000002202f518","ABjDJDJ1zmeZe6buX2OaYyTgwjCbL23gewBXpibZN3Fg4=")),
("求职面试高频题库","小红书","新增(候选)","面试不要太老实，10道高频题这样答更加分","时光呀","205","反直觉(别老实)+数字题库","反直觉+数字",url("6a3c8b2e000000001702bcf7","ABv-3lRYhIDm7h69PJmm-psBbHBsJLjTZ_lVB3o6WYhUo=")),
("求职面试高频题库","小红书","新增(候选)","Google高频leetcode题目(近三个月)","Ashley","166","大厂技术题库+时效性标注","权威身份(Google)+时效",url("6a02dc6e00000000070224a5","ABkXFKZTxfjr6_TrB3sY0sOA-K5taMsyVE9QSW6hablDg=")),
("求职面试高频题库","小红书","新增(候选)","dataanalyst面试不背题真的寸步难行","Data数据疯狂找工","727","垂类岗位+背题必要性论证","垂类身份+断言",url("6779c9d10000000014023238","AB9cpYF-ZWzIveHeA2AtzN98WAacJRjF2RJ5lQSJcV-zk=")),
("求职面试高频题库","小红书","新增(候选)","一图看懂job interview面试题+答案","欣启跃产品经理冲刺班","126","图解形式的题库整理","图解体+可抄答案",url("68278c9200000000230022b9","AB0MxZw7BZ5AiI4sn2LwckawnhJqiF68WolzJjglY3_P4=")),
("求职面试高频题库","小红书","新增(候选)","一个能过🇺🇸四大药厂面试的方法（附题）","Chance","184","垂类行业(药厂)方法论+附题","垂类身份+方法论",url("6a3e428400000000220085b5","ABxm3-cF08tTEhqkuZg6DPPCynxW8zp6fJYlA2-wa2jWE=")),

("高情商破局","小红书","新增(候选)","老祖宗专门留下来对付小人的心理博弈术","周树人","731","传统智慧包装的职场博弈术","权威身份(老祖宗)+博弈术",url("6a2cde2d000000000e021800","ABvv6wVJ00WKzUkhM1Ny1eSJU-L4-dEA-j3QPwVRqCVbs=")),
("高情商破局","小红书","新增(候选)","对付小人的唯一心法：不斗人，只破局","安燃成长日记","—","反直觉心法(不斗人)+破局概念","反直觉断言",url("6a0dee580000000036000610","ABPcuE_PJ-3g8HlVeSNJWQrFDmzR8fUJjD8TgxvWbzbrk=")),
("高情商破局","小红书","新增(候选)","最好的反击，是摧毁问题（第二期）","会说话的豆腐","1.5万","反击哲学观点+系列号","反直觉+系列号",url("6a166f250000000037036204","AByMxq8t6-mPiSc09PYVZrXaIAhlijVratlwnInYkeDzY=")),
("高情商破局","小红书","新增(候选)","草台班子里最聪明的不是往上爬","女帝思维Yi Xuan","524","职场生存智慧+俚语标签(草台班子)","反直觉+俚语标签",url("6a2b4f83000000001700adfe","ABrNMG-fFQfpTgVuJ6fpDpdl4MbUMmN6_9ii6VSD2z0NY=")),
("高情商破局","小红书","新增(候选)","《毛选》：别人甩锅，一招破局","思维高阶认知","—","经典理论+具体职场场景(甩锅)","权威典籍+场景痛点",url("6a252f680000000022019ae9","ABoGqkTEi5uSr-Aru_6YamWuXI6vzDf-i8ShZItg_xaWI=")),
("高情商破局","小红书","新增(候选)","让你level飙升的8个“破局”思维","清心读书成长","195","数字化清单体+职级提升承诺","数字+利益直给(level)",url("68bc37ca000000001d038d1a","ABPf5z01JGrD5vbLCjbpnpB0QivUgGIYZmj5LjFcRxpI4=")),
]

for row in content_rows:
    ws1.append(row)

for r in range(2, ws1.max_row+1):
    for c in range(1, len(headers1)+1):
        ws1.cell(row=r, column=c).font = BODY_FONT
        ws1.cell(row=r, column=c).alignment = WRAP

style_header(ws1, len(headers1))
autosize(ws1, [16,8,10,32,16,8,30,20,42])

# ============ Sheet 2: 根词表现表 ============
ws2 = wb.create_sheet("根词表现表")
headers2 = ["根词","类型","本次抓取(估)","本次新增(过滤后)","命中率","平均热度","说明","升级/退休"]
ws2.append(headers2)

perf_rows = [
("职场表达","种子(同日第3轮复查)",20,6,"30.0%","约12995(被7.1万拉高)","种子永久保留,同日第3轮命中率低于run1/run2","无变化"),
("面试技巧","种子(同日第3轮复查)",20,1,"5.0%","659","种子永久保留,供给已被前两轮充分收割","无变化"),
("面试什么话该说","种子(同日第3轮复查)",20,0,"0%","—","20条全部重复,完全饱和","无变化"),
("面试什么话不能说","种子(同日第3轮复查)",20,3,"15.0%","约4109","顶部命中自家账号笔记","无变化"),
("面试反杀面试官","活跃(长空档复查,距07-19隔25天)",33,10,"30.3%","约7640","Mr Jonathan系账号贡献过半新增,长空档规律验证","无变化(继续活跃)"),
("面试被hr阴阳怪气","活跃(长空档复查,距07-19隔25天)",20,12,"60.0%","约2648","本轮全场最高命中率,长空档规律强验证","无变化(继续活跃)"),
("求职面试高频题库","候选(首投,08-12新收割)",20,7,"35.0%","约282","未达40%阈值,维持候选待复投","维持候选"),
("高情商破局","候选(首投,08-12新收割)",20,6,"30.0%(原始80%经过滤)","约4113","与已活跃职场破局话术大全/职场话术反制技巧中度重叠","维持候选"),
]
for row in perf_rows:
    ws2.append(row)

for r in range(2, ws2.max_row+1):
    for c in range(1, len(headers2)+1):
        ws2.cell(row=r, column=c).font = BODY_FONT
        ws2.cell(row=r, column=c).alignment = WRAP
style_header(ws2, len(headers2))
autosize(ws2, [20,26,14,16,10,26,34,18])

# ============ Sheet 3: 词库验证表 ============
ws3 = wb.create_sheet("词库验证表")
headers3 = ["验证长句","场景类型","条件a(下拉补全)","条件b(结果页数量级)","条件c(前排10篇赞藏水平)","判定结果","意图强度/竞争密度","发现的答案空缺","收割的新词清单"]
ws3.append(headers3)

verify_rows = [
("社招offer可以先接了再拒吗","事件","通过(下拉出现近义'社招offer可以先接了再拒嘛'等)","较多(AI总结77篇笔记)","不通过,多篇远超500赞(4万/2.4万/2.2万/5321/5189等)","放弃","—","谈offer接拒赛道已饱和(负向发现)","社招offer可以先接了再拒嘛/邮件点了接受offer可以不去吗等(未单独收割,饱和方向不追加)"),
("口头offer到正式offer一般要多久","事件","通过(下拉直接出现'口头offer到正式offer要多久时间'等多条近乎一致补全)","较多(25条可见样本)","通过,前排绝大多数<500赞(113/42/15/94/30/67/9/121/45/133等)","已验证(意图强度=高,竞争密度=低)","高/低","供给多为个人晒结果贴,缺标准化时间线整理+超时催问话术模板;与已验证'社招一面多久出结果''社招oc一般多久发offer'构成offer流程时间线焦虑簇","口头offer后怎么催进度/邮件点了接受offer可以不去吗"),
("无领导小组讨论的题目有哪些","事件","通过(下拉出现'无领导小组讨论的经典题目'等)","较多(25条可见样本)","不通过,多篇远超500赞(9326/1.4万/4318/2117等)","放弃(与已升级活跃的'无领导小组模拟题库'高度重叠)","—","题库类内容已饱和,与关键词池路径'无领导小组模拟题库'(100%命中)重叠","两难式无领导小组面试题目/无领导小组材料分析题(转入关键词池)"),
("控场能力和领导能力怎么培养","情境","不通过(搜索框无任何该词或近似词补全)","较多(25条可见样本)","不通过,多篇远超500赞(1.6万/1.7万/2.7万/1.1万/9685等)","放弃(两项条件均不通过)","—","控场/领导气场内容已饱和","领导气场培养(转入关键词池)"),
("第一次面试紧张怎么办","症状","通过(下拉出现'第一次面试有点紧张怎么办'等多条近乎一致补全)","较多(25条可见样本)","不通过,多篇远超500赞(4.5万/8.9万/2.1万/1.8万/7424等)","放弃(与已升级活跃的'面试紧张怎么快速缓解'高度重叠)","—","面试紧张赛道已饱和,与关键词池路径'面试紧张怎么快速缓解'(70%命中)重叠","—(未额外收割,避免与已活跃词重复)"),
]
for row in verify_rows:
    ws3.append(row)

for r in range(2, ws3.max_row+1):
    for c in range(1, len(headers3)+1):
        ws3.cell(row=r, column=c).font = BODY_FONT
        ws3.cell(row=r, column=c).alignment = WRAP
style_header(ws3, len(headers3))
autosize(ws3, [24,10,26,18,30,20,14,34,28])

# 评论区原话补充sheet (bonus, small)
ws4 = wb.create_sheet("评论区原话")
headers4 = ["日期","来源链接","用户原话","暴露的处境","候选词"]
ws4.append(headers4)
comment_rows = [
("2026-08-13","https://www.xiaohongshu.com/search_result/68006d90000000001e0060e4?xsec_token=AB52wtiI_BSQNzpsbLbNL-qw0A80yz9m55XGQi3tgCmdo=&xsec_source=","对我也有同样的体验 国内的面试 面试官没有给予我尊重","面试中感觉不被尊重/得不到基本礼貌对待","面试官不尊重人怎么办"),
("2026-08-13","https://www.xiaohongshu.com/search_result/68006d90000000001e0060e4?xsec_token=AB52wtiI_BSQNzpsbLbNL-qw0A80yz9m55XGQi3tgCmdo=&xsec_source=","要我国内时间下午面试🙃逆天","跨时区/工作时间被要求配合不合理的面试安排","面试时间安排不合理怎么办"),
]
for row in comment_rows:
    ws4.append(row)
for r in range(2, ws4.max_row+1):
    for c in range(1, len(headers4)+1):
        ws4.cell(row=r, column=c).font = BODY_FONT
        ws4.cell(row=r, column=c).alignment = WRAP
style_header(ws4, len(headers4))
autosize(ws4, [12,50,40,34,26])

wb.save("职场表达与面试技巧_2026-08-13_run3.xlsx")
print("saved")
